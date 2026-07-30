from __future__ import annotations

from copy import deepcopy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import importlib.util
import threading
import unittest
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import load_workbook
import pandas as pd

from backend.app.main import (
    _filter_forecast_sources,
    _require_output_run,
    download_forecast_output_csv,
    download_forecast_output_excel,
    download_forecast_output_pdf,
    get_forecast_output_run_preview,
)
import pio_platform.output_center as output_center


class OutputCenterTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        months = pd.period_range("2024-01", periods=30, freq="M")
        rows = []
        for index, month in enumerate(months):
            for brand, anchor, model, plc, part, base in (
                ("H", "HMA", "Elantra", "Floor Mat", "H-MAT", 100.0),
                ("K", "KUS", "Sportage", "Cargo Tray", "K-TRAY", 70.0),
            ):
                quantity = base + index
                rows.append(
                    {
                        "month": str(month),
                        "brand": brand,
                        "anchorBrand": anchor,
                        "entityKey": f"{anchor}::{model.upper()}",
                        "modelName": model,
                        "plc": plc,
                        "partNumber": part,
                        "partDescription": part,
                        "lifecycleStatus": "Active",
                        "lifecycleStatusCode": "active",
                        "installationQuantity": quantity,
                        "pioRevenue": quantity * 80.0,
                    }
                )
        facts = pd.DataFrame(rows)
        wholesale = pd.DataFrame(
            [
                {
                    "month": str(month),
                    "brand": anchor,
                    "anchorBrand": anchor,
                    "modelName": model,
                    "modelKey": model.upper(),
                    "wholesaleUnits": base + index,
                }
                for index, month in enumerate(months)
                for anchor, model, base in (
                    ("HMA", "Elantra", 200.0),
                    ("KUS", "Sportage", 150.0),
                )
            ]
        )
        cls.working_days = pd.DataFrame(
            {
                "month": months.astype(str),
                "workingDays": [21.0] * len(months),
            }
        )
        cls.filtered_facts, cls.filtered_wholesale = _filter_forecast_sources(
            facts,
            wholesale,
            brand=["HMA"],
            model=[],
            part=[],
            start_date="2024-01-01",
            end_date="2026-06-30",
        )
        output_center.clear_output_run_cache()
        forecast_builder = output_center.build_forecast_center
        planning_builder = output_center.build_part_planning_bundle
        with patch.object(
            output_center,
            "build_forecast_center",
            wraps=forecast_builder,
        ) as forecast_spy, patch.object(
            output_center,
            "build_part_planning_bundle",
            wraps=planning_builder,
        ) as planning_spy:
            cls.output_run, cls.first_reused = (
                output_center.create_or_get_output_run(
                    workbook_id="workbook-1",
                    sheet_name="PIO_Sales_Data",
                    source_filename="source.xlsx",
                    source_hash="source-hash-1",
                    facts=cls.filtered_facts,
                    working_days=cls.working_days,
                    wholesale_long=cls.filtered_wholesale,
                    filters={
                        "brand": ["HMA"],
                        "model": [],
                        "part": [],
                        "startDate": "2024-01-01",
                        "endDate": "2026-06-30",
                    },
                    horizon=3,
                    top_n=10,
                    use_working_days=True,
                    use_seasonality=True,
                    tariff_impact_pct=0.0,
                    min_monthly_volume=5.0,
                    requested_strategy="naive_last",
                    latest_sales_month_is_complete=True,
                    latest_sales_date=pd.Timestamp("2026-06-30"),
                )
            )
            cls.reused_run, cls.second_reused = (
                output_center.create_or_get_output_run(
                    workbook_id="workbook-1",
                    sheet_name="PIO_Sales_Data",
                    source_filename="source.xlsx",
                    source_hash="source-hash-1",
                    facts=cls.filtered_facts,
                    working_days=cls.working_days,
                    wholesale_long=cls.filtered_wholesale,
                    filters={
                        "brand": ["HMA"],
                        "model": [],
                        "part": [],
                        "startDate": "2024-01-01",
                        "endDate": "2026-06-30",
                    },
                    horizon=3,
                    top_n=10,
                    use_working_days=True,
                    use_seasonality=True,
                    tariff_impact_pct=0.0,
                    min_monthly_volume=5.0,
                    requested_strategy="naive_last",
                    latest_sales_month_is_complete=True,
                    latest_sales_date=pd.Timestamp("2026-06-30"),
                )
            )
            cls.forecast_build_count = forecast_spy.call_count
            cls.part_planning_build_count = planning_spy.call_count

    def _base_run_kwargs(self) -> dict:
        return {
            "workbook_id": "workbook-1",
            "sheet_name": "PIO_Sales_Data",
            "source_filename": "source.xlsx",
            "source_hash": "source-hash-1",
            "facts": self.filtered_facts,
            "working_days": self.working_days,
            "wholesale_long": self.filtered_wholesale,
            "filters": {
                "brand": ["HMA"],
                "model": [],
                "part": [],
                "startDate": "2024-01-01",
                "endDate": "2026-06-30",
            },
            "horizon": 3,
            "top_n": 10,
            "use_working_days": True,
            "use_seasonality": True,
            "tariff_impact_pct": 0.0,
            "min_monthly_volume": 5.0,
            "requested_strategy": "naive_last",
            "latest_sales_month_is_complete": True,
            "latest_sales_date": pd.Timestamp("2026-06-30"),
        }

    def _fake_output_run(self, **kwargs) -> dict:
        metadata = {
            **self.output_run["metadata"],
            "runId": kwargs["run_id"],
            "workbookId": kwargs["workbook_id"],
            "sheetName": kwargs["sheet_name"],
            "sourceHash": kwargs["source_hash"],
            "sourceSignature": kwargs["source_signature"],
            "sourceFilename": kwargs["source_filename"],
            "filters": kwargs["normalized_filters"],
            **kwargs["settings"],
        }
        return {
            "metadata": metadata,
            "executiveSummary": {
                **self.output_run["executiveSummary"],
                "metadata": metadata,
            },
            "payloads": self.output_run["payloads"],
            "partQuantity": self.output_run["partQuantity"],
            "partRevenue": self.output_run["partRevenue"],
            "artifacts": self.output_run["artifacts"],
        }

    def _create_fake_cached_run(self, **overrides):
        kwargs = {**self._base_run_kwargs(), **overrides}
        with patch.object(
            output_center,
            "_build_output_run",
            side_effect=self._fake_output_run,
        ):
            return output_center.create_or_get_output_run(**kwargs)

    def test_same_parameters_reuse_complete_immutable_run(self) -> None:
        self.assertFalse(self.first_reused)
        self.assertTrue(self.second_reused)
        self.assertEqual(
            self.output_run["metadata"]["runId"],
            self.reused_run["metadata"]["runId"],
        )
        self.assertEqual(self.forecast_build_count, 3)
        self.assertEqual(self.part_planning_build_count, 1)
        self.assertEqual(
            set(self.output_run["payloads"]),
            {"revenue", "quantity", "wholesale_quantity"},
        )
        self.assertEqual(
            {record["brand"] for record in self.output_run["payloads"]["revenue"]["brandRecords"]},
            {"HMA"},
        )
        self.assertEqual(self.output_run["metadata"]["filters"]["brand"], ["HMA"])

    def test_preview_and_excel_share_canonical_totals_and_metadata(self) -> None:
        preview = output_center.output_run_preview(self.output_run)
        workbook = load_workbook(
            BytesIO(self.output_run["artifacts"]["detailedExcel"]),
            data_only=False,
        )
        self.assertEqual(
            workbook.sheetnames,
            [
                "Executive_Summary",
                "Revenue_Forecast",
                "Quantity_Forecast",
                "Part_Planning",
                "Wholesale_Drivers",
                "Model_Performance",
                "QA_Assumptions",
                "Run_Metadata",
            ],
        )
        self.assertEqual(workbook["Run_Metadata"].sheet_state, "hidden")
        self.assertTrue(str(workbook["QA_Assumptions"]["B5"].value).startswith("="))
        self.assertTrue(str(workbook["Part_Planning"]["O6"].value).startswith("="))
        summary_sheet = workbook["Executive_Summary"]
        self.assertEqual(summary_sheet["B5"].value, preview["metadata"]["runId"])
        self.assertEqual(summary_sheet["B7"].value, preview["metadata"]["sourceHash"])
        self.assertEqual(summary_sheet["B8"].value, preview["metadata"]["cutoff"])
        for offset, expected in enumerate(
            preview["executiveSummary"]["headlineTotals"],
            start=19,
        ):
            self.assertEqual(summary_sheet.cell(offset, 1).value, expected["month"])
            self.assertEqual(summary_sheet.cell(offset, 2).value, expected["periodType"])
            self.assertAlmostEqual(
                float(summary_sheet.cell(offset, 3).value),
                float(expected["revenue"]),
            )
            self.assertAlmostEqual(
                float(summary_sheet.cell(offset, 4).value),
                float(expected["quantity"]),
            )
            self.assertAlmostEqual(
                float(summary_sheet.cell(offset, 5).value),
                float(expected["wholesale_quantity"]),
            )
        self.assertTrue(
            all(
                check["status"] == "PASS"
                for check in preview["executiveSummary"]["reconciliation"].values()
            )
        )

    def test_current_view_csv_reads_run_without_forecast_rebuild(self) -> None:
        with patch.object(
            output_center,
            "build_forecast_center",
            wraps=output_center.build_forecast_center,
        ) as forecast_spy:
            csv_text = output_center.build_current_view_csv(
                self.output_run,
                metric="revenue",
                level="model",
            )
            preview = output_center.output_run_preview(self.output_run)
            _ = self.output_run["artifacts"]["detailedExcel"]
        self.assertEqual(forecast_spy.call_count, 0)
        header = csv_text.splitlines()[0]
        for field in (
            "runId",
            "sourceHash",
            "cutoff",
            "requestedStrategy",
            "effectiveStrategy",
            "period",
            "periodType",
            "unit",
        ):
            self.assertIn(field, header)
        self.assertIn(preview["metadata"]["runId"], csv_text)
        self.assertIn(preview["metadata"]["sourceHash"], csv_text)

    def test_identity_changes_for_every_governed_source_filter_and_setting(self) -> None:
        output_center.clear_output_run_cache()
        base_kwargs = self._base_run_kwargs()
        changed_facts = self.filtered_facts.copy()
        changed_facts.loc[changed_facts.index[0], "pioRevenue"] += 1.0
        changed_working_days = self.working_days.copy()
        changed_working_days.loc[changed_working_days.index[0], "workingDays"] += 1.0
        changed_wholesale = self.filtered_wholesale.copy()
        changed_wholesale.loc[
            changed_wholesale.index[0], "wholesaleUnits"
        ] += 1.0
        variations = [
            {"facts": changed_facts},
            {"working_days": changed_working_days},
            {"wholesale_long": changed_wholesale},
            {
                "filters": {
                    **base_kwargs["filters"],
                    "model": ["Elantra"],
                }
            },
            {"horizon": 2},
            {"top_n": 9},
            {"use_working_days": False},
            {"use_seasonality": False},
            {"tariff_impact_pct": 1.0},
            {"min_monthly_volume": 6.0},
            {"requested_strategy": "mean"},
        ]
        with patch.object(
            output_center,
            "_build_output_run",
            side_effect=self._fake_output_run,
        ):
            base_run, _ = output_center.create_or_get_output_run(**base_kwargs)
            changed_ids = {
                output_center.create_or_get_output_run(
                    **{**base_kwargs, **variation}
                )[0]["metadata"]["runId"]
                for variation in variations
            }
        self.assertEqual(len(changed_ids), len(variations))
        self.assertNotIn(base_run["metadata"]["runId"], changed_ids)

    def test_ttl_expiry_returns_404_and_never_recomputes(self) -> None:
        output_center.clear_output_run_cache()
        clock = [100.0]
        with patch.object(
            output_center,
            "_monotonic",
            side_effect=lambda: clock[0],
        ), patch.object(
            output_center,
            "_build_output_run",
            side_effect=self._fake_output_run,
        ) as builder:
            output_run, _ = output_center.create_or_get_output_run(
                **self._base_run_kwargs()
            )
            clock[0] += output_center.OUTPUT_RUN_TTL_SECONDS + 1
            with self.assertRaises(HTTPException) as context:
                _require_output_run(
                    "workbook-1",
                    "PIO_Sales_Data",
                    output_run["metadata"]["runId"],
                )
        self.assertEqual(context.exception.status_code, 404)
        self.assertEqual(builder.call_count, 1)

    def test_lru_count_and_total_byte_limits_evict_oldest_runs(self) -> None:
        base_kwargs = self._base_run_kwargs()
        output_center.clear_output_run_cache()
        with patch.object(output_center, "OUTPUT_RUN_CACHE_LIMIT", 2), patch.object(
            output_center,
            "OUTPUT_RUN_CACHE_BYTES_LIMIT",
            10_000,
        ), patch.object(
            output_center,
            "_estimate_output_run_bytes",
            return_value=10,
        ), patch.object(
            output_center,
            "_build_output_run",
            side_effect=self._fake_output_run,
        ):
            runs = [
                output_center.create_or_get_output_run(
                    **{**base_kwargs, "top_n": top_n}
                )[0]
                for top_n in (8, 9, 10)
            ]
        self.assertEqual(output_center.output_run_cache_stats()["count"], 2)
        self.assertIsNone(
            output_center.get_output_run(
                runs[0]["metadata"]["runId"],
                workbook_id="workbook-1",
                sheet_name="PIO_Sales_Data",
            )
        )

        output_center.clear_output_run_cache()
        with patch.object(output_center, "OUTPUT_RUN_CACHE_LIMIT", 8), patch.object(
            output_center,
            "OUTPUT_RUN_CACHE_BYTES_LIMIT",
            100,
        ), patch.object(
            output_center,
            "_estimate_output_run_bytes",
            return_value=60,
        ), patch.object(
            output_center,
            "_build_output_run",
            side_effect=self._fake_output_run,
        ):
            byte_runs = [
                output_center.create_or_get_output_run(
                    **{**base_kwargs, "top_n": top_n}
                )[0]
                for top_n in (9, 10)
            ]
        stats = output_center.output_run_cache_stats()
        self.assertEqual(stats, {"count": 1, "bytes": 60, "inFlight": 0})
        self.assertIsNone(
            output_center.get_output_run(
                byte_runs[0]["metadata"]["runId"],
                workbook_id="workbook-1",
                sheet_name="PIO_Sales_Data",
            )
        )
        self.assertIsNotNone(
            output_center.get_output_run(
                byte_runs[1]["metadata"]["runId"],
                workbook_id="workbook-1",
                sheet_name="PIO_Sales_Data",
            )
        )

    def test_same_identity_is_single_flight(self) -> None:
        output_center.clear_output_run_cache()
        started = threading.Event()
        release = threading.Event()

        def blocking_build(**kwargs):
            started.set()
            self.assertTrue(release.wait(timeout=5))
            return self._fake_output_run(**kwargs)

        with patch.object(
            output_center,
            "_build_output_run",
            side_effect=blocking_build,
        ) as builder, ThreadPoolExecutor(max_workers=2) as executor:
            first = executor.submit(
                output_center.create_or_get_output_run,
                **self._base_run_kwargs(),
            )
            self.assertTrue(started.wait(timeout=5))
            second = executor.submit(
                output_center.create_or_get_output_run,
                **self._base_run_kwargs(),
            )
            self.assertEqual(output_center.output_run_cache_stats()["inFlight"], 1)
            release.set()
            first_result = first.result(timeout=5)
            second_result = second.result(timeout=5)
        self.assertEqual(builder.call_count, 1)
        self.assertEqual(
            first_result[0]["metadata"]["runId"],
            second_result[0]["metadata"]["runId"],
        )
        self.assertEqual({first_result[1], second_result[1]}, {False, True})

    def test_different_identities_build_without_global_serialization(self) -> None:
        output_center.clear_output_run_cache()
        both_started = threading.Barrier(2)

        def barrier_build(**kwargs):
            both_started.wait(timeout=5)
            return self._fake_output_run(**kwargs)

        with patch.object(
            output_center,
            "_build_output_run",
            side_effect=barrier_build,
        ) as builder, ThreadPoolExecutor(max_workers=2) as executor:
            first = executor.submit(
                output_center.create_or_get_output_run,
                **{**self._base_run_kwargs(), "top_n": 9},
            )
            second = executor.submit(
                output_center.create_or_get_output_run,
                **{**self._base_run_kwargs(), "top_n": 10},
            )
            first_run = first.result(timeout=5)[0]
            second_run = second.result(timeout=5)[0]
        self.assertEqual(builder.call_count, 2)
        self.assertNotEqual(
            first_run["metadata"]["runId"],
            second_run["metadata"]["runId"],
        )

    def test_wrong_workbook_or_sheet_is_404(self) -> None:
        output_center.clear_output_run_cache()
        output_run, _ = self._create_fake_cached_run()
        for workbook_id, sheet_name in (
            ("wrong-workbook", "PIO_Sales_Data"),
            ("workbook-1", "Wrong_Sheet"),
        ):
            with self.assertRaises(HTTPException) as context:
                _require_output_run(
                    workbook_id,
                    sheet_name,
                    output_run["metadata"]["runId"],
                )
            self.assertEqual(context.exception.status_code, 404)

    def test_all_http_run_reads_never_call_forecast_builder(self) -> None:
        output_center.clear_output_run_cache()
        output_run, _ = self._create_fake_cached_run()
        run_id = output_run["metadata"]["runId"]
        with patch.object(
            output_center,
            "build_forecast_center",
            side_effect=AssertionError("download attempted forecast rebuild"),
        ) as builder:
            preview = get_forecast_output_run_preview(
                "workbook-1",
                "PIO_Sales_Data",
                run_id,
            )
            excel = download_forecast_output_excel(
                "workbook-1",
                "PIO_Sales_Data",
                run_id,
            )
            pdf = download_forecast_output_pdf(
                "workbook-1",
                "PIO_Sales_Data",
                run_id,
            )
            csv = download_forecast_output_csv(
                "workbook-1",
                "PIO_Sales_Data",
                run_id,
                metric="revenue",
                level="brand",
            )
        self.assertEqual(builder.call_count, 0)
        self.assertEqual(preview["metadata"]["runId"], run_id)
        self.assertEqual(excel.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(pdf.media_type, "application/pdf")
        self.assertEqual(csv.media_type, "text/csv")

    def test_missing_or_expired_run_is_404_without_recompute(self) -> None:
        with self.assertRaises(HTTPException) as context:
            _require_output_run(
                "workbook-1",
                "PIO_Sales_Data",
                "pio-output-run-v1__missing",
            )
        self.assertEqual(context.exception.status_code, 404)
        self.assertIn("never silently recompute", str(context.exception.detail))

    def test_reference_portfolio_is_revenue_only_for_output_run(self) -> None:
        output_center.clear_output_run_cache()
        payloads = self.output_run["payloads"]

        def fake_forecast(*args, **kwargs):
            return deepcopy(payloads[str(kwargs["metric"])])

        with patch.object(
            output_center,
            "build_forecast_center",
            side_effect=fake_forecast,
        ) as forecast_spy, patch.object(
            output_center,
            "build_part_planning_bundle",
            return_value=(
                deepcopy(self.output_run["partQuantity"]),
                deepcopy(self.output_run["partRevenue"]),
            ),
        ):
            reference_run, _ = output_center.create_or_get_output_run(
                workbook_id="workbook-1",
                sheet_name="PIO_Sales_Data",
                source_filename="source.xlsx",
                source_hash="source-hash-1",
                facts=self.filtered_facts,
                working_days=self.working_days,
                wholesale_long=self.filtered_wholesale,
                filters={
                    "brand": ["HMA"],
                    "model": [],
                    "part": [],
                    "startDate": "2024-01-01",
                    "endDate": "2026-06-30",
                },
                horizon=3,
                top_n=10,
                use_working_days=True,
                use_seasonality=True,
                tariff_impact_pct=0.0,
                min_monthly_volume=5.0,
                requested_strategy="reference_portfolio",
                latest_sales_month_is_complete=True,
                latest_sales_date=pd.Timestamp("2026-06-30"),
            )
        self.assertEqual(
            [call.kwargs["model_strategy"] for call in forecast_spy.call_args_list],
            ["reference_portfolio", "auto", "auto"],
        )
        self.assertEqual(
            reference_run["metadata"]["effectiveStrategies"],
            {
                "revenue": "reference_portfolio",
                "quantity": "auto",
                "wholesale_quantity": "auto",
            },
        )

    @unittest.skipUnless(
        importlib.util.find_spec("reportlab") is not None,
        "reportlab is approved but not installed in this workspace",
    )
    def test_pdf_contains_canonical_metadata_periods_and_totals(self) -> None:
        pdf_bytes = output_center.build_executive_summary_pdf(
            self.output_run["executiveSummary"]
        )
        self.assertTrue(pdf_bytes.startswith(b"%PDF-"))
        pdf_text = pdf_bytes.decode("latin-1", errors="ignore")
        metadata = self.output_run["metadata"]
        for value in (
            metadata["runId"],
            metadata["sourceHash"],
            metadata["cutoff"],
            metadata["requestedStrategy"],
            *metadata["forecastPeriods"],
        ):
            self.assertIn(str(value), pdf_text)
        for row in self.output_run["executiveSummary"]["headlineTotals"]:
            self.assertIn(f"{float(row['revenue']):.2f}", pdf_text)
            self.assertIn(f"{float(row['quantity']):.2f}", pdf_text)
            self.assertIn(f"{float(row['wholesale_quantity']):.2f}", pdf_text)


if __name__ == "__main__":
    unittest.main()
