from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from pio_platform.forecast_center import (
    _allocation_accuracy_diagnostics,
    _allocation_accuracy_for_level,
    _forecast_exceptions,
    _registered_evidence_gate,
    build_forecast_center,
    build_part_planning_records,
    clear_forecast_diagnostic_caches,
)
from pio_platform.ets_experiments import VALIDATED_REFERENCE_PORTFOLIO
from pio_platform.ml_challengers import ML_CHALLENGER_IDS
from pio_platform.sop_workbook import build_sop_workbook_bytes


class ForecastCenterTests(unittest.TestCase):
    def setUp(self) -> None:
        months = pd.date_range("2023-01-01", periods=43, freq="MS")
        rows = []
        for index, month in enumerate(months):
            partial = 0.55 if month == months[-1] else 1.0
            rows.extend(
                [
                    {
                        "month": month.strftime("%Y-%m"),
                        "brand": "H",
                        "entityKey": "H::ELANTRA",
                        "modelName": "Elantra",
                        "plc": "Floor Mat",
                        "partNumber": "H-MAT",
                        "partDescription": "Carpet Floor Mat",
                        "lifecycleStatus": "Active",
                        "installationQuantity": (100 + index) * partial,
                        "pioRevenue": (100 + index) * 80 * partial,
                    },
                    {
                        "month": month.strftime("%Y-%m"),
                        "brand": "H",
                        "entityKey": "H::ELANTRA",
                        "modelName": "Elantra",
                        "plc": "Cargo Tray",
                        "partNumber": "H-TRAY",
                        "partDescription": "Cargo Tray",
                        "lifecycleStatus": "Active",
                        "installationQuantity": (40 + index / 2) * partial,
                        "pioRevenue": (40 + index / 2) * 110 * partial,
                    },
                    {
                        "month": month.strftime("%Y-%m"),
                        "brand": "K",
                        "entityKey": "K::SPORTAGE",
                        "modelName": "Sportage",
                        "plc": "Floor Mat",
                        "partNumber": "K-MAT",
                        "partDescription": "Carpet Floor Mat",
                        "lifecycleStatus": "Active",
                        "installationQuantity": (70 + index / 3) * partial,
                        "pioRevenue": (70 + index / 3) * 75 * partial,
                    },
                ]
            )
        self.facts = pd.DataFrame(rows)
        self.working_days = pd.DataFrame(
            {
                "month": months.strftime("%Y-%m"),
                "workingDays": [20 + (index % 3) for index in range(len(months))],
            }
        )
        self.wholesale = pd.DataFrame(
            {
                "month": list(months.strftime("%Y-%m")) * 3,
                "brand": ["HMA"] * len(months) + ["GMA"] * len(months) + ["KUS"] * len(months),
                "modelName": ["Elantra"] * len(months) + ["GV70"] * len(months) + ["Sportage"] * len(months),
                "modelKey": ["ELANTRA"] * len(months) + ["GV70"] * len(months) + ["SPORTAGE"] * len(months),
                "wholesaleUnits": [200 + index for index in range(len(months))]
                + [50 + index for index in range(len(months))]
                + [150 + index for index in range(len(months))],
            }
        )

    def test_revenue_nowcast_and_hierarchy_reconcile(self) -> None:
        payload = build_forecast_center(
            self.facts,
            self.working_days,
            self.wholesale,
            metric="revenue",
            level="model_plc",
            horizon=3,
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
            include_all_records=True,
        )
        self.assertEqual(payload["summary"]["nowcastMonths"], ["2026-07"])
        self.assertEqual(payload["summary"]["pureForecastMonths"], ["2026-08", "2026-09"])
        self.assertEqual(payload["summary"]["reconciliation"]["status"], "PASS")
        self.assertLessEqual(len(payload["topAccessories"]), 10)
        self.assertTrue(all(record["plc"] in {"Floor Mat", "Cargo Tray"} for record in payload["records"]))
        self.assertEqual({record["brand"] for record in payload["brandRecords"]}, {"HMA", "KUS"})
        self.assertEqual(
            payload["summary"]["accuracyScope"]["evaluatedGrain"],
            "month × selected official brand anchor(s) from HMA/GMA/KUS",
        )
        self.assertIn("day-of-month cutoff is not a working-day count", payload["summary"]["periodExplanation"])
        self.assertIn(
            "current business rule",
            payload["summary"]["anchorPolicy"]["denominatorPolicy"]["HMA"],
        )
        self.assertEqual(
            {item["evaluationScopeId"] for item in payload["summary"]["evaluationScopes"]},
            {
                "application_recent_h1",
                "governed_h123_24m_h1_18_h2_17_h3_16_official_total_51",
            },
        )
        for record in payload["brandRecords"]:
            for forecast in record["forecast"]:
                self.assertGreaterEqual(forecast["lower"], 0.0)
                self.assertLessEqual(forecast["lower"], forecast["point"])
                self.assertLessEqual(forecast["point"], forecast["upper"])
        self.assertTrue(
            all(
                forecast["validationStatus"] == "unvalidated_child_interval_coverage"
                and forecast["empiricalCoverage"] is None
                for record in payload["modelRecords"]
                for forecast in record["forecast"]
            )
        )
        self.assertTrue(
            all(
                interval["validationStatus"]
                in {"unvalidated_partial_anchor_scope", "unvalidated_nowcast"}
                for interval in payload["summary"]["predictionIntervals"]["officialTotal"]
            )
        )
        unit_revenue = {
            (record["modelName"], record["plc"]): record["expectedUnitRevenue"]
            for record in payload["modelPlcRecords"]
            if record["allocationRoute"] == "regular_allocation"
        }
        self.assertAlmostEqual(unit_revenue[("Elantra", "Floor Mat")], 80.0, places=6)
        self.assertAlmostEqual(unit_revenue[("Elantra", "Cargo Tray")], 110.0, places=6)
        self.assertAlmostEqual(unit_revenue[("Sportage", "Floor Mat")], 75.0, places=6)
        formulas = {item["name"]: item for item in payload["summary"]["formulaCatalog"]}
        self.assertIn("SumOfPIS_CRP_CFM_PRI", formulas["Monthly target actual"]["formula"])
        self.assertIn("recent-6", formulas["Exact-part unit revenue"]["formula"])

    def test_pretrained_ml_challengers_apply_july_mtd_after_offline_training(self) -> None:
        with patch(
            "pio_platform.forecast_center._allocation_accuracy_diagnostics",
            return_value={},
        ), patch(
            "pio_platform.forecast_center._forecast_exceptions",
            return_value=[],
        ):
            for model_id in sorted(ML_CHALLENGER_IDS):
                payload = build_forecast_center(
                    self.facts,
                    self.working_days,
                    self.wholesale,
                    metric="revenue",
                    level="brand",
                    horizon=3,
                    model_strategy=model_id,
                    latest_sales_month_is_complete=False,
                    latest_sales_date=pd.Timestamp("2026-07-28"),
                    source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
                )
                self.assertEqual(payload["summary"]["nowcastMonths"], ["2026-07"])
                self.assertEqual(payload["summary"]["pureForecastMonths"], ["2026-08", "2026-09"])
                self.assertEqual(payload["summary"]["factors"]["modelStrategy"], model_id)
                for record in payload["brandRecords"]:
                    july = record["forecast"][0]
                    self.assertEqual(july["month"], "2026-07")
                    self.assertEqual(july["forecastType"], "Nowcast")
                    self.assertGreater(july["actualToDate"], 0.0)
                    self.assertGreaterEqual(july["point"], july["actualToDate"])
                    self.assertGreater(july["statisticalBaseline"], 0.0)
                    self.assertEqual(record["requestedModelStrategy"], model_id)
                    self.assertIn("No training occurred", record["selectionNote"])

    def test_wholesale_keeps_hma_gma_and_kus_as_separate_anchors(self) -> None:
        payload = build_forecast_center(
            self.facts,
            self.working_days,
            self.wholesale,
            metric="wholesale_quantity",
            level="model",
            horizon=2,
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
        )
        self.assertEqual(
            {record["brand"] for record in payload["brandRecords"]},
            {"HMA", "GMA", "KUS"},
        )
        self.assertEqual(payload["summary"]["reconciliation"]["status"], "PASS")
        self.assertEqual(payload["summary"]["forecastMonths"][0], "2026-07")
        self.assertEqual(payload["summary"]["nowcastMonths"], ["2026-07"])

    def test_wholesale_models_stopped_before_current_year_are_zero(self) -> None:
        stopped = pd.DataFrame(
            {
                "month": ["2023-01", "2025-07", "2023-01", "2025-10"],
                "brand": ["KUS", "KUS", "HMA", "HMA"],
                "modelName": ["Stinger", "Stinger", "Nexo", "Nexo"],
                "modelKey": ["STINGER", "STINGER", "NEXO", "NEXO"],
                "wholesaleUnits": [40.0, 30.0, 25.0, 20.0],
            }
        )
        wholesale = pd.concat([self.wholesale, stopped], ignore_index=True)

        payload = build_forecast_center(
            self.facts,
            self.working_days,
            wholesale,
            metric="wholesale_quantity",
            level="model",
            horizon=2,
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
            include_all_records=True,
        )
        stopped_records = {
            record["modelName"]: record
            for record in payload["modelRecords"]
            if record["modelName"] in {"Nexo", "Stinger"}
        }

        self.assertEqual(stopped_records["Nexo"]["allocationRoute"], "excluded_lifecycle")
        self.assertEqual(stopped_records["Stinger"]["allocationRoute"], "excluded_lifecycle")
        self.assertTrue(
            all(
                forecast["value"] == 0
                for record in stopped_records.values()
                for forecast in record["forecast"]
            )
        )
        self.assertEqual(payload["summary"]["reconciliation"]["status"], "PASS")

    def test_reference_portfolio_is_revenue_only_and_brand_specific(self) -> None:
        payload = build_forecast_center(
            self.facts,
            self.working_days,
            self.wholesale,
            metric="revenue",
            level="brand",
            horizon=2,
            model_strategy="reference_portfolio",
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
        )
        methods = {
            record["brand"]: record["selectedModel"]
            for record in payload["brandRecords"]
        }

        self.assertTrue(methods["HMA"].startswith("hw_add_add__heuristic"))
        self.assertEqual(methods["KUS"], "working_day_adjusted_seasonal")
        residual_methods = {
            record["brand"]: {
                row["backtest_model"]
                for row in record["rollingOriginResiduals"]
            }
            for record in payload["brandRecords"]
        }
        self.assertTrue(
            all(method.startswith("hw_add_add__heuristic") for method in residual_methods["HMA"])
        )
        self.assertEqual(residual_methods["KUS"], {"working_day_adjusted_seasonal"})
        self.assertTrue(
            all(
                record["forecast"][0]["validationStatus"] == "unvalidated_nowcast"
                and record["forecast"][0]["empiricalCoverage"] is None
                for record in payload["brandRecords"]
            )
        )
        self.assertTrue(
            all(
                record["forecast"][1]["validationStatus"]
                == "validated_brand_rolling_origin"
                for record in payload["brandRecords"]
            )
        )
        self.assertEqual(
            payload["summary"]["modelGovernance"]["requestedStrategy"],
            "reference_portfolio",
        )
        self.assertEqual(
            payload["summary"]["modelGovernance"]["referenceMethodStatus"],
            "validated_implementation_applied_to_unvalidated_source",
        )
        with self.assertRaisesRegex(ValueError, "only for Revenue"):
            build_forecast_center(
                self.facts,
                self.working_days,
                self.wholesale,
                metric="quantity",
                level="brand",
                model_strategy="reference_portfolio",
            )

    def test_exact_parts_allocate_to_plc_parent(self) -> None:
        payload = build_forecast_center(
            self.facts,
            self.working_days,
            self.wholesale,
            metric="quantity",
            level="model_plc",
            horizon=2,
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
            include_all_records=True,
        )
        part_records = build_part_planning_records(
            self.facts,
            payload["modelPlcRecords"],
            metric="quantity",
            latest_complete_month=payload["summary"]["latestCompleteMonth"],
        )
        self.assertGreater(len(part_records), 0)
        parent = {
            (record["brand"], record["entityKey"], record["plc"], item["month"]): item["value"]
            for record in payload["modelPlcRecords"]
            for item in record["forecast"]
        }
        children: dict[tuple[str, str, str, str], float] = {}
        for record in part_records:
            key = (record["brand"], record["entityKey"], record["plc"], record["month"])
            children[key] = children.get(key, 0.0) + record["value"]
        for key, value in parent.items():
            self.assertAlmostEqual(children[key], value, places=6)

    def test_low_volume_and_new_models_are_routed_without_breaking_reconciliation(self) -> None:
        facts = self.facts.copy()
        low = facts[facts["modelName"] == "Elantra"].copy()
        low["entityKey"] = "H::LOW VOLUME"
        low["modelName"] = "Low Volume"
        low["installationQuantity"] = 0.1
        low["pioRevenue"] = 8.0
        low["lifecycleStatus"] = "Active"
        low["lifecycleStatusCode"] = "active"
        new = facts[facts["modelName"] == "Elantra"].tail(8).copy()
        new["entityKey"] = "H::NEW EV"
        new["modelName"] = "New EV"
        new["installationQuantity"] = 30.0
        new["pioRevenue"] = 2400.0
        new["lifecycleStatus"] = "New / limited history"
        new["lifecycleStatusCode"] = "new"
        routed = pd.concat([facts, low, new], ignore_index=True)

        payload = build_forecast_center(
            routed,
            self.working_days,
            self.wholesale,
            metric="quantity",
            level="model",
            horizon=2,
            latest_sales_month_is_complete=False,
            latest_sales_date=pd.Timestamp("2026-07-22"),
            include_all_records=True,
        )
        routes = {record["modelName"]: record["allocationRoute"] for record in payload["modelRecords"]}
        self.assertEqual(routes["Low Volume"], "excluded_low_volume")
        self.assertEqual(routes["New EV"], "new_model_proxy")
        self.assertEqual(payload["summary"]["reconciliation"]["status"], "PASS")

    def test_sop_workbook_has_governed_sheets_and_formulas(self) -> None:
        payloads = {}
        for metric in ["revenue", "quantity", "wholesale_quantity"]:
            payloads[metric] = build_forecast_center(
                self.facts,
                self.working_days,
                self.wholesale,
                metric=metric,
                level="brand",
                horizon=2,
                latest_sales_month_is_complete=False,
                latest_sales_date=pd.Timestamp("2026-07-22"),
                include_all_records=True,
            )
        part_quantity = build_part_planning_records(
            self.facts,
            payloads["quantity"]["modelPlcRecords"],
            metric="quantity",
            latest_complete_month=payloads["quantity"]["summary"]["latestCompleteMonth"],
        )
        part_revenue = build_part_planning_records(
            self.facts,
            payloads["revenue"]["modelPlcRecords"],
            metric="revenue",
            latest_complete_month=payloads["revenue"]["summary"]["latestCompleteMonth"],
        )
        output = build_sop_workbook_bytes(
            source_filename="source.xlsx",
            revenue=payloads["revenue"],
            quantity=payloads["quantity"],
            wholesale=payloads["wholesale_quantity"],
            part_quantity=part_quantity,
            part_revenue=part_revenue,
            working_days=self.working_days.to_dict("records"),
        )
        workbook = load_workbook(BytesIO(output), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Executive_Summary",
                "Revenue_Forecast",
                "Quantity_Forecast",
                "Part_Planning",
                "Wholesale_Drivers",
                "Model_Performance",
                "QA_Assumptions",
            ],
        )
        self.assertTrue(str(workbook["QA_Assumptions"]["B5"].value).startswith("="))
        self.assertTrue(str(workbook["Part_Planning"]["O6"].value).startswith("="))

    def test_registered_metrics_gate_rejects_filtered_scope_with_same_hash(self) -> None:
        brands = [
            {"brand": anchor}
            for anchor in ("HMA", "GMA", "KUS")
        ]
        eligible = _registered_evidence_gate(
            metric="revenue",
            model_strategy="reference_portfolio",
            source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
            latest_complete_month="2026-06",
            brand_records=brands,
            evaluation_scope_eligible=True,
            evaluation_scope_metadata={
                "filtersApplied": False,
                "requestCutoff": "2026-06",
                "target": "revenue",
            },
        )
        filtered = _registered_evidence_gate(
            metric="revenue",
            model_strategy="reference_portfolio",
            source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
            latest_complete_month="2026-06",
            brand_records=brands,
            evaluation_scope_eligible=False,
            evaluation_scope_metadata={
                "filtersApplied": True,
                "requestCutoff": "2026-06",
                "target": "revenue",
            },
        )
        self.assertTrue(eligible["eligible"])
        self.assertFalse(filtered["eligible"])
        self.assertEqual(
            eligible["evaluationScopeId"],
            "governed_h123_24m_h1_18_h2_17_h3_16_official_total_51",
        )
        complete_metadata = dict(eligible["requestScope"])
        for field in (
            "filtersApplied",
            "requestCutoff",
            "target",
        ):
            incomplete = dict(complete_metadata)
            incomplete.pop(field)
            decision = _registered_evidence_gate(
                metric="revenue",
                model_strategy="reference_portfolio",
                source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
                latest_complete_month="2026-06",
                brand_records=brands,
                evaluation_scope_eligible=True,
                evaluation_scope_metadata=incomplete,
            )
            self.assertFalse(decision["eligible"], field)
        trusted = dict(VALIDATED_REFERENCE_PORTFOLIO["evaluationEvidence"])
        for field in ("foldKeyIdentity", "expectedFoldCounts", "predictionCoverage"):
            invalid_trusted = dict(trusted)
            invalid_trusted.pop(field)
            self_reported = {
                **complete_metadata,
                "foldKeyIdentity": trusted["foldKeyIdentity"],
                "foldCounts": trusted["expectedFoldCounts"],
                "predictionCoverage": 1.0,
                "fullCoverage": True,
            }
            decision = _registered_evidence_gate(
                metric="revenue",
                model_strategy="reference_portfolio",
                source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
                latest_complete_month="2026-06",
                brand_records=brands,
                evaluation_scope_eligible=True,
                evaluation_scope_metadata=self_reported,
                trusted_evidence=invalid_trusted,
            )
            self.assertFalse(decision["eligible"], field)
        wrong_identity = dict(trusted)
        wrong_identity["foldKeyIdentity"] = "caller-supplied-but-wrong"
        self.assertFalse(
            _registered_evidence_gate(
                metric="revenue",
                model_strategy="reference_portfolio",
                source_hash=VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
                latest_complete_month="2026-06",
                brand_records=brands,
                evaluation_scope_eligible=True,
                evaluation_scope_metadata=complete_metadata,
                trusted_evidence=wrong_identity,
            )["eligible"]
        )
        mismatch_cases = [
            {"metric": "quantity"},
            {"model_strategy": "auto"},
            {"source_hash": "wrong"},
            {"latest_complete_month": "2026-05"},
            {"brand_records": [{"brand": "HMA"}, {"brand": "KUS"}]},
        ]
        defaults = {
            "metric": "revenue",
            "model_strategy": "reference_portfolio",
            "source_hash": VALIDATED_REFERENCE_PORTFOLIO["sourceHash"],
            "latest_complete_month": "2026-06",
            "brand_records": brands,
            "evaluation_scope_eligible": True,
            "evaluation_scope_metadata": complete_metadata,
        }
        for mismatch in mismatch_cases:
            decision = _registered_evidence_gate(**{**defaults, **mismatch})
            self.assertFalse(decision["eligible"], mismatch)

    def test_three_allocation_only_levels_use_complete_51_fold_contract(self) -> None:
        months = pd.period_range("2023-01", periods=42, freq="M")
        rows = []
        for index, month in enumerate(months):
            for model_name, plc, part, share in (
                ("Elantra", "Floor Mat", "A", 0.6),
                ("Elantra", "Cargo Tray", "B", 0.4),
                ("Sonata", "Floor Mat", "C", 0.5),
                ("Sonata", "Cargo Tray", "D", 0.5),
            ):
                quantity = (100.0 + index) * share
                rows.append(
                    {
                        "month": str(month),
                        "brand": "HMA",
                        "entityKey": f"HMA::{model_name.upper()}",
                        "modelName": model_name,
                        "plc": plc,
                        "partNumber": part,
                        "installationQuantity": quantity,
                        "pioRevenue": quantity * (80.0 if plc == "Floor Mat" else 100.0),
                    }
                )
        diagnostics = _allocation_accuracy_diagnostics(
            pd.DataFrame(rows),
            metric="revenue",
            latest_complete_month="2026-06",
            source_hash="test",
        )
        self.assertEqual({item["level"] for item in diagnostics}, {"Model", "PLC", "PIS_PNO"})
        self.assertTrue(all(item["scope"] == "allocationOnly" for item in diagnostics))
        self.assertTrue(all(item["foldCount"] == 51 for item in diagnostics))
        self.assertTrue(all(item["coverage"] == 1.0 for item in diagnostics))
        self.assertTrue(all(item["validationStatus"] == "validated_allocation_only" for item in diagnostics))

    def test_allocation_origin_signal_is_leakage_safe_and_zero_cases_degrade(self) -> None:
        months = pd.period_range("2023-01", periods=42, freq="M")
        base = pd.DataFrame(
            [
                {
                    "month": str(month),
                    "brand": "HMA",
                    "entityKey": "HMA::A",
                    "modelName": "A",
                    "installationQuantity": 0.0 if index == 40 else 10.0,
                    "pioRevenue": 0.0 if index == 40 else 800.0,
                }
                for index, month in enumerate(months)
            ]
        )
        original = _allocation_accuracy_for_level(
            base,
            metric="quantity",
            level="Model",
            child_dimensions=["brand", "entityKey", "modelName"],
            parent_dimensions=["brand"],
            latest_complete_month="2026-06",
            source_hash="test",
        )
        perturbed = base.copy()
        perturbed.loc[perturbed["month"] == "2026-06", "installationQuantity"] = 9999.0
        changed = _allocation_accuracy_for_level(
            perturbed,
            metric="quantity",
            level="Model",
            child_dimensions=["brand", "entityKey", "modelName"],
            parent_dimensions=["brand"],
            latest_complete_month="2026-06",
            source_hash="test",
        )
        original_fold = next(
            item
            for item in original["foldAudits"]
            if item["originMonth"] == "2026-05"
            and item["targetMonth"] == "2026-06"
            and item["horizon"] == 1
        )
        changed_fold = next(
            item
            for item in changed["foldAudits"]
            if item["originMonth"] == "2026-05"
            and item["targetMonth"] == "2026-06"
            and item["horizon"] == 1
        )
        self.assertEqual(
            original_fold["originSignalHash"],
            changed_fold["originSignalHash"],
        )
        self.assertGreater(original["zeroParentActualCount"], 0)

        zero_signal = base.copy()
        zero_signal["installationQuantity"] = 0.0
        zero_signal.loc[zero_signal["month"] == "2025-01", "installationQuantity"] = 10.0
        degraded = _allocation_accuracy_for_level(
            zero_signal,
            metric="quantity",
            level="Model",
            child_dimensions=["brand", "entityKey", "modelName"],
            parent_dimensions=["brand"],
            latest_complete_month="2026-06",
            source_hash="test",
        )
        self.assertGreater(degraded["zeroSignalParentCount"], 0)
        self.assertLess(degraded["predictionCoverage"], 1.0)
        self.assertNotEqual(degraded["validationStatus"], "validated_allocation_only")

    def test_forecast_exceptions_cover_all_governed_reason_codes(self) -> None:
        months = pd.period_range("2025-07", periods=12, freq="M")
        source_rows = [
                {
                    "month": str(month),
                    "brand": "HMA",
                    "entityKey": "HMA::GAP",
                    "modelName": "Gap model",
                    "plc": "Floor Mat",
                    "partNumber": "GAP-PART",
                    "lifecycleStatus": "Inactive",
                    "lifecycleStatusCode": "inactive",
                    "installationQuantity": (
                        10.0 if index in {0, 2, 4} else 0.0
                    ),
                    "pioRevenue": (
                        800.0 if index in {0, 2, 4} else 0.0
                    ),
                }
                for index, month in enumerate(months)
                if index in {0, 2, 4}
            ]
        source_rows.append(
            {
                "month": "2026-06",
                "brand": "HMA",
                "entityKey": "HMA::NEW",
                "modelName": "New model",
                "plc": "Cargo Tray",
                "partNumber": "NEW-PART",
                "lifecycleStatus": "New / limited history",
                "lifecycleStatusCode": "new",
                "installationQuantity": 10.0,
                "pioRevenue": 800.0,
            }
        )
        source = pd.DataFrame(source_rows)
        base = {
            "brand": "HMA",
            "brandName": "Hyundai Motor America",
            "entityKey": "HMA::GAP",
            "modelName": "Gap model",
            "seriesKey": "HMA::GAP",
            "plc": "",
            "activeMonths": 3,
            "historyMonths": 12,
            "monthlyAverage": 2.5,
            "historyVolume": 30.0,
            "expectedUnitRevenue": 80.0,
            "forecast": [{"month": "2026-07", "value": 0.0}],
        }
        model_records = [
            {
                **base,
                "allocationRoute": "excluded_lifecycle",
                "lifecycleStatus": "Inactive",
                "lifecycleStatusCode": "inactive",
            },
            {
                **base,
                "seriesKey": "HMA::NEW",
                "entityKey": "HMA::NEW",
                "modelName": "New model",
                "allocationRoute": "new_model_proxy",
                "lifecycleStatus": "New / limited history",
                "lifecycleStatusCode": "new",
                "monthlyAverage": 10.0,
                "forecast": [{"month": "2026-07", "value": 10.0}],
            },
            {
                **base,
                "seriesKey": "HMA::PLANNER",
                "entityKey": "HMA::PLANNER",
                "modelName": "Planner review residual",
                "allocationRoute": "planner_review_residual",
                "monthlyAverage": 10.0,
                "activeMonths": 12,
            },
        ]
        exceptions = _forecast_exceptions(
            source,
            model_records,
            [],
            [],
            metric="revenue",
            latest_complete_month="2026-06",
            min_monthly_volume=5.0,
        )
        self.assertEqual(
            {item["reasonCode"] for item in exceptions},
            {
                "inactive_discontinued",
                "low_volume",
                "insufficient_active_months",
                "history_gaps",
                "recent_zero_streak",
                "new_reintroduced_proxy",
                "zero_forecast_historical_unit_price",
                "planner_review_residual",
            },
        )
        self.assertTrue(
            all(
                {"severity", "scope", "grain", "seriesKey", "forecastMonth", "evidence", "suggestedAction"}
                .issubset(item)
                for item in exceptions
            )
        )

    def test_strategy_residuals_and_diagnostics_are_cached_by_scope(self) -> None:
        import pio_platform.forecast_center as forecast_center_module

        clear_forecast_diagnostic_caches()
        wrapped = forecast_center_module.rolling_origin_residuals
        signature_builder = forecast_center_module._diagnostic_source_signatures
        allocation_builder = forecast_center_module._preaggregate_allocation_monthly
        common = {
            "metric": "revenue",
            "horizon": 2,
            "model_strategy": "naive_last",
            "latest_sales_month_is_complete": False,
            "latest_sales_date": pd.Timestamp("2026-07-22"),
        }
        with patch.object(
            forecast_center_module,
            "rolling_origin_residuals",
            wraps=wrapped,
        ) as residual_builder, patch.object(
            forecast_center_module,
            "_diagnostic_source_signatures",
            wraps=signature_builder,
        ) as signature_spy, patch.object(
            forecast_center_module,
            "_preaggregate_allocation_monthly",
            wraps=allocation_builder,
        ) as allocation_spy:
            first = build_forecast_center(
                self.facts,
                self.working_days,
                self.wholesale,
                level="brand",
                **common,
            )
            first_calls = residual_builder.call_count
            self.assertEqual(signature_spy.call_count, 1)
            self.assertEqual(allocation_spy.call_count, 1)
            second = build_forecast_center(
                self.facts,
                self.working_days,
                self.wholesale,
                level="model",
                **common,
            )
            self.assertEqual(signature_spy.call_count, 2)
            self.assertEqual(allocation_spy.call_count, 1)
        self.assertGreater(first_calls, 0)
        self.assertEqual(residual_builder.call_count, first_calls)
        self.assertFalse(first["summary"]["diagnosticCache"]["residualCacheHit"])
        self.assertFalse(first["summary"]["diagnosticCache"]["governanceCacheHit"])
        self.assertTrue(second["summary"]["diagnosticCache"]["residualCacheHit"])
        self.assertTrue(second["summary"]["diagnosticCache"]["governanceCacheHit"])
        self.assertTrue(
            all(
                row["backtest_model"] == "naive_last"
                for record in second["brandRecords"]
                for row in record["rollingOriginResiduals"]
            )
        )

    def test_auto_residuals_use_fold_local_selector_not_naive_fallback(self) -> None:
        import pio_platform.hierarchical_forecasting as hierarchy_module

        months = pd.date_range("2024-01-01", periods=30, freq="MS")
        series = pd.Series(range(100, 130), index=months, dtype=float)

        def selected(history, working_days, **kwargs):
            horizon = int(kwargs["horizon"])
            return {
                "model": f"fold_selected_{len(history)}",
                "forecast": [float(history.iloc[-1])] * horizon,
            }

        with patch.object(
            hierarchy_module,
            "_select_forecast_candidate",
            side_effect=selected,
        ) as selector:
            rows = hierarchy_module.rolling_origin_residuals(
                series,
                {},
                entity="HMA",
                use_working_days=True,
                use_seasonality=True,
                model_strategy="auto",
            )
        self.assertGreater(selector.call_count, 0)
        self.assertTrue(
            all(str(row["backtest_model"]).startswith("fold_selected_") for row in rows)
        )
        self.assertNotIn("naive_last", {row["backtest_model"] for row in rows})

    def test_exact_part_recent_six_zero_routes_parent_to_planner_review(self) -> None:
        months = pd.period_range("2024-01", periods=30, freq="M")
        facts = pd.DataFrame(
            [
                {
                    "month": str(month),
                    "brand": "H",
                    "entityKey": "H::ELANTRA",
                    "modelName": "Elantra",
                    "plc": "Old PLC",
                    "partNumber": "OLD-PART",
                    "partDescription": "Old part",
                    "lifecycleStatus": "Active",
                    "lifecycleStatusCode": "active",
                    "installationQuantity": 10.0 if index < 21 else 0.0,
                    "pioRevenue": 800.0 if index < 21 else 0.0,
                }
                for index, month in enumerate(months)
            ]
        )
        working_days = pd.DataFrame(
            {"month": months.astype(str), "workingDays": [21.0] * len(months)}
        )
        clear_forecast_diagnostic_caches()
        payload = build_forecast_center(
            facts,
            working_days,
            self.wholesale,
            metric="revenue",
            level="brand",
            horizon=3,
            model_strategy="seasonal_naive",
            latest_sales_month_is_complete=True,
            latest_sales_date=pd.Timestamp("2026-06-30"),
        )
        exact_part_exceptions = [
            item
            for item in payload["forecastExceptions"]
            if item["scope"] == "PIS_PNO"
        ]
        forecast_months = set(payload["summary"]["forecastMonths"])
        self.assertEqual(
            {
                item["forecastMonth"]
                for item in exact_part_exceptions
                if item["reasonCode"] == "planner_review_residual"
            },
            forecast_months,
        )
        self.assertEqual(
            {
                item["forecastMonth"]
                for item in exact_part_exceptions
                if item["reasonCode"] == "zero_forecast_historical_unit_price"
            },
            forecast_months,
        )

    def test_integrated_exceptions_are_level_invariant_and_month_truthful(self) -> None:
        months = pd.period_range("2024-01", periods=30, freq="M")
        rows = []
        for month in months:
            for part_number in ("LOW-A", "LOW-B", "LOW-C"):
                rows.append(
                    {
                        "month": str(month),
                        "brand": "H",
                        "entityKey": "H::ELANTRA",
                        "modelName": "Elantra",
                        "plc": "Planner PLC",
                        "partNumber": part_number,
                        "partDescription": part_number,
                        "lifecycleStatus": "Active",
                        "lifecycleStatusCode": "active",
                        "installationQuantity": 2.0,
                        "pioRevenue": 160.0,
                    }
                )
        for index in (0, 2, 4):
            rows.append(
                {
                    "month": str(months[index]),
                    "brand": "H",
                    "entityKey": "H::INACTIVE",
                    "modelName": "Inactive model",
                    "plc": "Gap PLC",
                    "partNumber": "GAP-PART",
                    "partDescription": "Gap part",
                    "lifecycleStatus": "Inactive",
                    "lifecycleStatusCode": "inactive",
                    "installationQuantity": 1.0,
                    "pioRevenue": 80.0,
                }
            )
        rows.append(
            {
                "month": str(months[-1]),
                "brand": "H",
                "entityKey": "H::NEW",
                "modelName": "New model",
                "plc": "New PLC",
                "partNumber": "NEW-PART",
                "partDescription": "New part",
                "lifecycleStatus": "New / limited history",
                "lifecycleStatusCode": "new",
                "installationQuantity": 10.0,
                "pioRevenue": 800.0,
            }
        )
        facts = pd.DataFrame(rows)
        working_days = pd.DataFrame(
            {"month": months.astype(str), "workingDays": [21.0] * len(months)}
        )
        clear_forecast_diagnostic_caches()
        common = {
            "facts": facts,
            "working_days": working_days,
            "wholesale_long": self.wholesale,
            "metric": "revenue",
            "horizon": 3,
            "model_strategy": "naive_last",
            "latest_sales_month_is_complete": True,
            "latest_sales_date": pd.Timestamp("2026-06-30"),
        }
        brand_payload = build_forecast_center(level="brand", **common)
        model_payload = build_forecast_center(level="model", **common)
        brand_exceptions = brand_payload["forecastExceptions"]
        self.assertEqual(
            {item["exceptionId"] for item in brand_exceptions},
            {item["exceptionId"] for item in model_payload["forecastExceptions"]},
        )
        self.assertEqual(
            {item["reasonCode"] for item in brand_exceptions},
            {
                "inactive_discontinued",
                "low_volume",
                "insufficient_active_months",
                "history_gaps",
                "recent_zero_streak",
                "new_reintroduced_proxy",
                "zero_forecast_historical_unit_price",
                "planner_review_residual",
            },
        )
        self.assertTrue(any(item["scope"] == "PIS_PNO" for item in brand_exceptions))
        forecast_months = set(brand_payload["summary"]["forecastMonths"])
        for item in brand_exceptions:
            if item["seriesLevel"]:
                self.assertIsNone(item["forecastMonth"])
            else:
                self.assertIn(item["forecastMonth"], forecast_months)


if __name__ == "__main__":
    unittest.main()
