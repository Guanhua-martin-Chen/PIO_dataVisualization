from __future__ import annotations

from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
RED = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "1F1F1F"
THIN_GRAY = Side(style="thin", color="D9E2F3")


def build_sop_workbook_bytes(
    *,
    source_filename: str,
    revenue: dict[str, Any],
    quantity: dict[str, Any],
    wholesale: dict[str, Any],
    part_quantity: list[dict[str, Any]],
    part_revenue: list[dict[str, Any]],
    working_days: list[dict[str, Any]],
    executive_summary: dict[str, Any] | None = None,
    run_metadata: dict[str, Any] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    _build_executive_summary(
        workbook,
        source_filename,
        revenue,
        quantity,
        wholesale,
        executive_summary=executive_summary,
    )
    _build_forecast_sheet(workbook, "Revenue_Forecast", revenue, currency=True)
    _build_forecast_sheet(workbook, "Quantity_Forecast", quantity, currency=False)
    _build_part_planning(workbook, part_quantity, part_revenue)
    _build_forecast_sheet(workbook, "Wholesale_Drivers", wholesale, currency=False)
    _build_model_performance(workbook, revenue, quantity, wholesale)
    _build_qa_assumptions(workbook, source_filename, revenue, quantity, wholesale, working_days)
    if run_metadata is not None:
        _build_run_metadata(workbook, run_metadata)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_executive_summary(
    workbook: Workbook,
    source_filename: str,
    revenue: dict[str, Any],
    quantity: dict[str, Any],
    wholesale: dict[str, Any],
    *,
    executive_summary: dict[str, Any] | None = None,
) -> None:
    if executive_summary is not None:
        _build_governed_executive_summary(workbook, executive_summary)
        return
    sheet = workbook.create_sheet("Executive_Summary")
    _title_block(
        sheet,
        "PIO FORECAST - EXECUTIVE SUMMARY",
        "Standardized Brand -> Model -> PLC forecast with exact-part planning detail",
        end_column=10,
    )
    metadata = [
        ("Source workbook", source_filename),
        ("Sales data through", revenue["summary"].get("dataThrough")),
        ("Completed training through", revenue["summary"].get("latestCompleteMonth")),
        ("Forecast periods", ", ".join(revenue["summary"].get("forecastMonths", []))),
        ("Brand definition", revenue["summary"].get("brandDefinition")),
        (
            "Dealer-wholesale mapping coverage",
            f"{float(revenue['summary'].get('anchorPolicy', {}).get('dealerWholesaleQuantityCoveragePct', 0.0)):.1f}%",
        ),
    ]
    row = 5
    for label, value in metadata:
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    _style_label_value_block(sheet, 5, row - 1, 1, 6)

    sheet.cell(5, 8, "MODEL STATUS")
    sheet.cell(6, 8, "Revenue hierarchy")
    sheet.cell(6, 9, revenue["summary"]["reconciliation"]["status"])
    sheet.cell(7, 8, "Quantity hierarchy")
    sheet.cell(7, 9, quantity["summary"]["reconciliation"]["status"])
    sheet.cell(8, 8, "Wholesale hierarchy")
    sheet.cell(8, 9, wholesale["summary"]["reconciliation"]["status"])
    for cell in sheet["H"][4:8]:
        cell.font = Font(bold=True, color=NAVY)
    for row_number in range(6, 9):
        status_cell = sheet.cell(row_number, 9)
        status_cell.fill = PatternFill("solid", fgColor=GREEN if status_cell.value == "PASS" else RED)
        status_cell.font = Font(bold=True, color=TEXT)

    kpi_row = 12
    headers = ["Metric", "Next Period Total", "WAPE", "Accuracy", "Forecast Type"]
    sheet.append([])
    for column, value in enumerate(headers, start=1):
        sheet.cell(kpi_row, column, value)
    metrics = [
        ("Revenue", revenue, True),
        ("PIO Quantity", quantity, False),
        ("Wholesale Quantity", wholesale, False),
    ]
    for offset, (label, payload, currency) in enumerate(metrics, start=1):
        total = sum(float(record.get("nextForecast", 0.0)) for record in payload.get("brandRecords", []))
        forecast_type = (
            "Nowcast"
            if payload["summary"].get("nowcastMonths")
            else "Forecast"
        )
        values = [
            label,
            total,
            payload["summary"].get("weightedWape"),
            (payload["summary"].get("accuracyPct") or 0.0) / 100.0,
            forecast_type,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(kpi_row + offset, column, value)
        sheet.cell(kpi_row + offset, 2).number_format = "$#,##0" if currency else "#,##0"
        sheet.cell(kpi_row + offset, 3).number_format = "0.0%"
        sheet.cell(kpi_row + offset, 4).number_format = "0.0%"
    _format_table_block(sheet, kpi_row, kpi_row + len(metrics), 1, len(headers))

    top_row = 18
    sheet.cell(top_row, 1, "TOP 10 PLC - REVENUE FORECAST")
    sheet.merge_cells(start_row=top_row, start_column=1, end_row=top_row, end_column=8)
    _section_header(sheet, top_row, 1, 8)
    top_headers = ["Rank", "PLC", "Historical Revenue Share", *revenue["summary"].get("forecastMonths", [])]
    for column, value in enumerate(top_headers, start=1):
        sheet.cell(top_row + 1, column, value)
    for offset, record in enumerate(revenue.get("topAccessories", []), start=2):
        row_number = top_row + offset
        sheet.cell(row_number, 1, record.get("rank"))
        sheet.cell(row_number, 2, record.get("plc"))
        sheet.cell(row_number, 3, float(record.get("historyRevenueSharePct", 0.0)) / 100.0)
        sheet.cell(row_number, 3).number_format = "0.0%"
        forecast_map = {str(item["month"]): float(item["value"]) for item in record.get("forecast", [])}
        for index, month in enumerate(revenue["summary"].get("forecastMonths", []), start=4):
            sheet.cell(row_number, index, forecast_map.get(month, 0.0))
            sheet.cell(row_number, index).number_format = "$#,##0"
    _format_table_block(
        sheet,
        top_row + 1,
        top_row + 1 + len(revenue.get("topAccessories", [])),
        1,
        len(top_headers),
    )

    note_row = top_row + 14
    sheet.cell(note_row, 1, "Management note")
    sheet.cell(
        note_row + 1,
        1,
        "HMA, GMA, and KUS forecasts are official anchors. Model, PLC, and PIS_PNO values are governed allocations; excluded low-volume/lifecycle volume is never hidden and any remainder is labeled for planner review.",
    )
    sheet.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 2, end_column=10)
    sheet.cell(note_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(note_row + 1, 1).fill = PatternFill("solid", fgColor=YELLOW)

    _set_widths(sheet, {"A": 22, "B": 28, "C": 22, "D": 18, "E": 18, "F": 18, "G": 3, "H": 22, "I": 18, "J": 18})
    sheet.freeze_panes = "A12"
    sheet.sheet_view.showGridLines = False


def _build_governed_executive_summary(
    workbook: Workbook,
    executive_summary: dict[str, Any],
) -> None:
    metadata = executive_summary["metadata"]
    sheet = workbook.create_sheet("Executive_Summary")
    _title_block(
        sheet,
        "PIO FORECAST - EXECUTIVE SUMMARY",
        "Immutable governed output run shared by web preview, PDF, and detailed Excel",
        end_column=10,
    )
    metadata_rows = [
        ("Forecast run ID", metadata["runId"]),
        ("Source workbook", metadata["sourceFilename"]),
        ("Source hash", metadata["sourceHash"]),
        ("Completed training through", metadata["cutoff"]),
        ("Requested strategy", metadata["requestedStrategy"]),
        (
            "Effective strategies",
            ", ".join(
                f"{metric}={strategy}"
                for metric, strategy in metadata["effectiveStrategies"].items()
            ),
        ),
        ("Nowcast periods", ", ".join(metadata["nowcastPeriods"]) or "None"),
        ("Forecast periods", ", ".join(metadata["forecastPeriods"]) or "None"),
        ("Created at", metadata["createdAt"]),
        ("Output contract", metadata["contractVersion"]),
    ]
    for row, (label, value) in enumerate(metadata_rows, start=5):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    _style_label_value_block(sheet, 5, 4 + len(metadata_rows), 1, 6)

    sheet.cell(5, 8, "RECONCILIATION")
    sheet.cell(5, 8).font = Font(bold=True, color=NAVY)
    for row, metric in enumerate(
        ("revenue", "quantity", "wholesale_quantity"),
        start=6,
    ):
        check = executive_summary["reconciliation"][metric]
        sheet.cell(row, 8, metric.replace("_", " ").title())
        sheet.cell(row, 9, check.get("status"))
        sheet.cell(row, 9).fill = PatternFill(
            "solid",
            fgColor=GREEN if check.get("status") == "PASS" else RED,
        )
        sheet.cell(row, 9).font = Font(bold=True, color=TEXT)

    headline_row = 17
    sheet.cell(headline_row, 1, "GOVERNED HEADLINE TOTALS")
    sheet.merge_cells(
        start_row=headline_row,
        start_column=1,
        end_row=headline_row,
        end_column=8,
    )
    _section_header(sheet, headline_row, 1, 8)
    headers = [
        "Forecast Month",
        "Period Type",
        "Revenue (USD)",
        "PIO Quantity (installed accessory units)",
        "Wholesale Quantity (vehicles)",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(headline_row + 1, column, value)
    for offset, item in enumerate(
        executive_summary["headlineTotals"],
        start=2,
    ):
        row = headline_row + offset
        values = [
            item["month"],
            item["periodType"],
            float(item["revenue"]),
            float(item["quantity"]),
            float(item["wholesale_quantity"]),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        sheet.cell(row, 3).number_format = "$#,##0.00"
        sheet.cell(row, 4).number_format = "#,##0.00"
        sheet.cell(row, 5).number_format = "#,##0.00"
    headline_end = headline_row + 1 + len(executive_summary["headlineTotals"])
    _format_table_block(sheet, headline_row + 1, headline_end, 1, len(headers))

    top_row = headline_end + 3
    sheet.cell(top_row, 1, "TOP 10 PLC - REVENUE FORECAST")
    sheet.merge_cells(
        start_row=top_row,
        start_column=1,
        end_row=top_row,
        end_column=8,
    )
    _section_header(sheet, top_row, 1, 8)
    forecast_months = [
        item["month"] for item in executive_summary["headlineTotals"]
    ]
    top_headers = ["Rank", "PLC", "Historical Revenue Share", *forecast_months]
    for column, value in enumerate(top_headers, start=1):
        sheet.cell(top_row + 1, column, value)
    top_records = executive_summary["topPlcs"][:10]
    for offset, record in enumerate(top_records, start=2):
        row = top_row + offset
        sheet.cell(row, 1, record.get("rank"))
        sheet.cell(row, 2, record.get("plc"))
        sheet.cell(
            row,
            3,
            float(record.get("historyRevenueSharePct", 0.0)) / 100.0,
        )
        sheet.cell(row, 3).number_format = "0.0%"
        forecast_map = {
            str(item["month"]): float(item["value"])
            for item in record.get("forecast", [])
        }
        for column, month in enumerate(forecast_months, start=4):
            sheet.cell(row, column, forecast_map.get(month, 0.0))
            sheet.cell(row, column).number_format = "$#,##0.00"
    _format_table_block(
        sheet,
        top_row + 1,
        top_row + 1 + len(top_records),
        1,
        len(top_headers),
    )
    _set_widths(
        sheet,
        {
            "A": 30,
            "B": 34,
            "C": 24,
            "D": 26,
            "E": 25,
            "F": 20,
            "G": 3,
            "H": 24,
            "I": 18,
            "J": 18,
        },
    )
    sheet.freeze_panes = "A18"
    sheet.sheet_view.showGridLines = False


def _build_forecast_sheet(
    workbook: Workbook,
    sheet_name: str,
    payload: dict[str, Any],
    *,
    currency: bool,
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    label = payload["summary"].get("metricLabel", sheet_name.replace("_", " "))
    _title_block(
        sheet,
        f"{label.upper()} FORECAST",
        payload["summary"].get("periodExplanation", ""),
        end_column=15,
    )
    headers = [
        "Forecast Month",
        "Forecast Type",
        "Level",
        "Brand Code",
        "Brand",
        "Model",
        "PLC",
        "Official Forecast",
        "Parent Forecast",
        "Allocation Share",
        "Reconciliation Factor",
        "Method",
        "Expected Unit Revenue",
        "Brand-Anchor WAPE",
        "Allocation Route",
    ]
    header_row = 5
    for column, value in enumerate(headers, start=1):
        sheet.cell(header_row, column, value)

    records: list[dict[str, Any]] = []
    records.extend(payload.get("brandRecords", []))
    records.extend(payload.get("modelRecords", []))
    records.extend(payload.get("modelPlcRecords", []))
    if payload["summary"].get("metric") == "wholesale_quantity":
        records = [*payload.get("brandRecords", []), *payload.get("modelRecords", [])]
    row = header_row + 1
    for record in records:
        for forecast in record.get("forecast", []):
            values = [
                forecast.get("month"),
                forecast.get("forecastType", "Forecast"),
                record.get("level"),
                record.get("brand"),
                record.get("brandName"),
                record.get("modelName"),
                record.get("plc"),
                forecast.get("value", 0.0),
                forecast.get("parentForecast"),
                forecast.get("allocationShare"),
                forecast.get("reconciliationFactor"),
                record.get("selectedModel"),
                record.get("expectedUnitRevenue"),
                record.get("wape"),
                record.get("allocationRoute") or "official_anchor",
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row, column, value)
            sheet.cell(row, 8).number_format = "$#,##0" if currency else "#,##0"
            sheet.cell(row, 9).number_format = "$#,##0" if currency else "#,##0"
            sheet.cell(row, 10).number_format = "0.000%"
            sheet.cell(row, 11).number_format = "0.000"
            sheet.cell(row, 13).number_format = "$#,##0"
            sheet.cell(row, 14).number_format = "0.0%"
            row += 1
    _format_table_block(sheet, header_row, max(header_row, row - 1), 1, len(headers))
    for data_row in range(header_row + 1, row):
        if sheet.cell(data_row, 2).value == "Nowcast":
            for column in range(1, len(headers) + 1):
                sheet.cell(data_row, column).fill = PatternFill("solid", fgColor=YELLOW)
    _add_table(sheet, f"{sheet_name.replace('_', '')}Table", header_row, max(header_row + 1, row - 1), len(headers))
    _set_widths(
        sheet,
        {
            "A": 16, "B": 14, "C": 12, "D": 12, "E": 24, "F": 22, "G": 24,
            "H": 18, "I": 18, "J": 16, "K": 20, "L": 25, "M": 20, "N": 18, "O": 24,
        },
    )
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:O{max(header_row, row - 1)}"
    sheet.sheet_view.showGridLines = False


def _build_part_planning(
    workbook: Workbook,
    quantity_records: list[dict[str, Any]],
    revenue_records: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Part_Planning")
    _title_block(
        sheet,
        "PIS_PNO PART PLANNING",
        "Exact-part quantity and revenue allocated from reconciled Model x PLC forecasts",
        end_column=16,
    )
    headers = [
        "Forecast Month",
        "Forecast Type",
        "Brand Code",
        "Brand",
        "Model",
        "PLC",
        "PIS_PNO",
        "Part Description",
        "Quantity Forecast",
        "Revenue Forecast",
        "Expected Unit Revenue",
        "Quantity Allocation Share",
        "Revenue Allocation Share",
        "Allocation Basis Month",
        "Reference Revenue (Qty x Unit Revenue)",
        "Allocation Route",
    ]
    header_row = 5
    for column, value in enumerate(headers, start=1):
        sheet.cell(header_row, column, value)
    revenue_lookup = {
        _part_key(record): record
        for record in revenue_records
    }
    row = header_row + 1
    for quantity in quantity_records:
        revenue = revenue_lookup.get(_part_key(quantity), {})
        values = [
            quantity.get("month"),
            quantity.get("forecastType"),
            quantity.get("brand"),
            quantity.get("brandName"),
            quantity.get("modelName"),
            quantity.get("plc"),
            quantity.get("partNumber"),
            quantity.get("partDescription"),
            quantity.get("value", 0.0),
            revenue.get("value", 0.0),
            revenue.get("expectedUnitRevenue") or quantity.get("expectedUnitRevenue"),
            quantity.get("allocationShare"),
            revenue.get("allocationShare"),
            quantity.get("allocationBasisMonth"),
            None,
            quantity.get("allocationRoute") or revenue.get("allocationRoute"),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        sheet.cell(row, 9).number_format = "#,##0"
        sheet.cell(row, 10).number_format = "$#,##0"
        sheet.cell(row, 11).number_format = "$#,##0.00"
        sheet.cell(row, 12).number_format = "0.000%"
        sheet.cell(row, 13).number_format = "0.000%"
        sheet.cell(row, 15, f'=IF(OR(I{row}=0,K{row}=0),"",I{row}*K{row})')
        sheet.cell(row, 15).number_format = "$#,##0"
        row += 1
    _format_table_block(sheet, header_row, max(header_row, row - 1), 1, len(headers))
    for data_row in range(header_row + 1, row):
        if sheet.cell(data_row, 2).value == "Nowcast":
            for column in range(1, len(headers) + 1):
                sheet.cell(data_row, column).fill = PatternFill("solid", fgColor=YELLOW)
    _add_table(sheet, "PartPlanningTable", header_row, max(header_row + 1, row - 1), len(headers))
    _set_widths(
        sheet,
        {
            "A": 16, "B": 14, "C": 12, "D": 24, "E": 22, "F": 24, "G": 20, "H": 28,
            "I": 18, "J": 18, "K": 20, "L": 20, "M": 20, "N": 20, "O": 18,
            "P": 24,
        },
    )
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:P{max(header_row, row - 1)}"
    sheet.sheet_view.showGridLines = False


def _build_model_performance(
    workbook: Workbook,
    revenue: dict[str, Any],
    quantity: dict[str, Any],
    wholesale: dict[str, Any],
) -> None:
    sheet = workbook.create_sheet("Model_Performance")
    _title_block(
        sheet,
        "BRAND ANCHOR MODEL PERFORMANCE",
        "Rolling-origin model selection and independent holdout diagnostics",
        end_column=11,
    )
    headers = [
        "Metric",
        "Brand Code",
        "Brand",
        "Selected Model",
        "Backtest Model",
        "History Months",
        "WAPE",
        "Accuracy",
        "Bias",
        "Next Forecast",
        "Selection Note",
    ]
    header_row = 5
    for column, value in enumerate(headers, start=1):
        sheet.cell(header_row, column, value)
    row = header_row + 1
    for payload in [revenue, quantity, wholesale]:
        for record in payload.get("brandRecords", []):
            values = [
                payload["summary"].get("metricLabel"),
                record.get("brand"),
                record.get("brandName"),
                record.get("selectedModel"),
                record.get("backtestModel"),
                record.get("historyMonths"),
                record.get("wape"),
                (record.get("accuracyPct") or 0.0) / 100.0,
                record.get("bias"),
                record.get("nextForecast"),
                record.get("selectionNote"),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row, column, value)
            sheet.cell(row, 7).number_format = "0.0%"
            sheet.cell(row, 8).number_format = "0.0%"
            sheet.cell(row, 9).number_format = "0.0%"
            sheet.cell(row, 10).number_format = "$#,##0" if payload["summary"].get("metric") == "revenue" else "#,##0"
            row += 1
    _format_table_block(sheet, header_row, max(header_row, row - 1), 1, len(headers))
    _add_table(sheet, "ModelPerformanceTable", header_row, max(header_row + 1, row - 1), len(headers))
    _set_widths(sheet, {"A": 22, "B": 12, "C": 24, "D": 24, "E": 22, "F": 16, "G": 14, "H": 14, "I": 14, "J": 18, "K": 58})
    sheet.freeze_panes = "A6"
    sheet.sheet_view.showGridLines = False


def _build_qa_assumptions(
    workbook: Workbook,
    source_filename: str,
    revenue: dict[str, Any],
    quantity: dict[str, Any],
    wholesale: dict[str, Any],
    working_days: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("QA_Assumptions")
    _title_block(
        sheet,
        "QA, FORMULAS & ASSUMPTIONS",
        "Controls, source notes, forecast formulas, and uploaded working days",
        end_column=10,
    )
    sheet["A5"] = "MODEL STATUS"
    sheet["B5"] = '=IF(COUNTIF(B9:B11,"FAIL")=0,"PASS","FAIL")'
    sheet["A5"].font = Font(bold=True, color=WHITE)
    sheet["A5"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["B5"].font = Font(bold=True)
    sheet["B5"].fill = PatternFill("solid", fgColor=GREEN)

    headers = ["Check", "Status", "Max Abs Delta", "Tolerance", "Where to fix", "Notes"]
    header_row = 8
    for column, value in enumerate(headers, start=1):
        sheet.cell(header_row, column, value)
    checks = [
        ("Revenue: Brand -> Model -> PLC", revenue["summary"]["reconciliation"]),
        ("Quantity: Brand -> Model -> PLC", quantity["summary"]["reconciliation"]),
        ("Wholesale: Brand -> Model", wholesale["summary"]["reconciliation"]),
    ]
    for offset, (label, check) in enumerate(checks, start=1):
        row = header_row + offset
        delta = max(float(check.get("brandToModelMaxAbsDelta", 0.0)), float(check.get("modelToPlcMaxAbsDelta", 0.0)))
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, f'=IF(C{row}<=D{row},"PASS","FAIL")')
        sheet.cell(row, 3, delta)
        sheet.cell(row, 4, float(check.get("tolerance", 0.01)))
        sheet.cell(row, 5, "Forecast engine reconciliation")
        sheet.cell(row, 6, "Parent and children must agree for every forecast month.")
        sheet.cell(row, 3).number_format = "0.000000"
        sheet.cell(row, 4).number_format = "0.000000"
    _format_table_block(sheet, header_row, header_row + len(checks), 1, len(headers))

    formula_row = 14
    sheet.cell(formula_row, 1, "FORMULA CATALOG")
    sheet.merge_cells(start_row=formula_row, start_column=1, end_row=formula_row, end_column=6)
    _section_header(sheet, formula_row, 1, 6)
    formula_headers = ["Metric", "Step", "Formula", "Logic"]
    for column, value in enumerate(formula_headers, start=1):
        sheet.cell(formula_row + 1, column, value)
    row = formula_row + 2
    for payload in [revenue, quantity, wholesale]:
        for formula in payload["summary"].get("formulaCatalog", []):
            sheet.cell(row, 1, payload["summary"].get("metricLabel"))
            sheet.cell(row, 2, formula.get("name"))
            sheet.cell(row, 3, formula.get("formula"))
            sheet.cell(row, 4, formula.get("logic"))
            row += 1
    _format_table_block(sheet, formula_row + 1, row - 1, 1, 4)

    source_row = row + 2
    sheet.cell(source_row, 1, "SOURCE & GOVERNANCE")
    sheet.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=6)
    _section_header(sheet, source_row, 1, 6)
    source_items = [
        ("Source workbook", source_filename),
        ("Brand policy", revenue["summary"].get("brandDefinition")),
        (
            "Wholesale denominator policy",
            "Current runtime uses dealer/non-fleet wholesale. The approved KUS contract separates Wholesale and Carpet Floor Mat Fleet baskets from 2026-06; implementation and backtest are pending.",
        ),
        (
            "Lifecycle policy",
            "Stopped series = 0; low volume = excluded; new/reintroduced = recent run-rate proxy; residuals require planner review.",
        ),
        ("Accessory policy", "PLC is the governed forecast accessory; PIS_PNO is planning detail."),
        ("Revenue policy", "Brand is official; Model, PLC, and PIS_PNO are reconciled allocations."),
        ("Part reference formula", "Reference Revenue = PIS_PNO quantity forecast x recent exact-part unit revenue; the reconciled Revenue Forecast remains official."),
        ("Primary horizon", ", ".join(revenue["summary"].get("forecastMonths", []))),
    ]
    for offset, (label, value) in enumerate(source_items, start=1):
        sheet.cell(source_row + offset, 1, label)
        sheet.cell(source_row + offset, 2, value)
        sheet.merge_cells(start_row=source_row + offset, start_column=2, end_row=source_row + offset, end_column=6)

    sheet["H5"] = "UPLOADED WORKING DAYS"
    sheet["H6"] = "Month"
    sheet["I6"] = "Working Days"
    for index, record in enumerate(working_days, start=7):
        sheet.cell(index, 8, record.get("month"))
        sheet.cell(index, 9, record.get("workingDays"))
        sheet.cell(index, 9).number_format = "0"
    _format_table_block(sheet, 6, max(7, 6 + len(working_days)), 8, 9)

    _set_widths(sheet, {"A": 28, "B": 20, "C": 62, "D": 65, "E": 28, "F": 48, "G": 4, "H": 18, "I": 16, "J": 16})
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=10):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A8"
    sheet.sheet_view.showGridLines = False


def _build_run_metadata(
    workbook: Workbook,
    run_metadata: dict[str, Any],
) -> None:
    sheet = workbook.create_sheet("Run_Metadata")
    sheet.append(["Field", "Value"])
    for field, value in run_metadata.items():
        sheet.append(
            [
                field,
                (
                    json.dumps(
                        value,
                        sort_keys=True,
                        separators=(",", ":"),
                        ensure_ascii=False,
                        default=str,
                    )
                    if isinstance(value, (dict, list))
                    else value
                ),
            ]
        )
    _format_table_block(sheet, 1, sheet.max_row, 1, 2)
    _set_widths(sheet, {"A": 30, "B": 110})
    sheet.freeze_panes = "A2"
    sheet.sheet_state = "hidden"


def _part_key(record: dict[str, Any]) -> tuple[str, ...]:
    return (
        str(record.get("month", "")),
        str(record.get("brand", "")),
        str(record.get("entityKey", "")),
        str(record.get("plc", "")),
        str(record.get("partNumber", "")),
    )


def _title_block(sheet: Any, title: str, subtitle: str, *, end_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    sheet.cell(1, 1, title)
    sheet.cell(1, 1).font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 10
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
    sheet.cell(3, 1, subtitle)
    sheet.cell(3, 1).font = Font(name="Aptos", size=10, italic=True, color=NAVY)
    sheet.cell(3, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 32


def _section_header(sheet: Any, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center")


def _format_table_block(sheet: Any, header_row: int, end_row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(header_row, column)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    for row in range(header_row + 1, end_row + 1):
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=WHITE if row % 2 == 0 else PALE_BLUE)
            cell.font = Font(color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=THIN_GRAY)


def _style_label_value_block(sheet: Any, start_row: int, end_row: int, start_column: int, end_column: int) -> None:
    for row in range(start_row, end_row + 1):
        sheet.cell(row, start_column).font = Font(bold=True, color=NAVY)
        sheet.cell(row, start_column).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for column in range(start_column + 1, end_column + 1):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=WHITE)


def _add_table(sheet: Any, name: str, header_row: int, end_row: int, end_column: int) -> None:
    if end_row <= header_row:
        return
    reference = f"A{header_row}:{_column_letter(end_column)}{end_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _set_widths(sheet: Any, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _column_letter(number: int) -> str:
    value = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        value = chr(65 + remainder) + value
    return value
